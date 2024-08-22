import pandas as pd
import  joblib,  shutil
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill


shutil.copyfile("product_original_reviews.xlsx", "product_detected_reviews.xlsx")  
workbook = load_workbook("product_detected_reviews.xlsx")
ws=workbook['Sheet1']
ws['Q15'] = "LABELS"
ws['Q15'].font = Font(bold=True)

svm_model = joblib.load("svm_model.pkl")

last_row = ws.max_row

for row in range(last_row, 17, -1):     
        if ws[f"A{row}"].value is None:
            ws.delete_rows(row)



df = pd.read_excel("product_detected_reviews.xlsx", header=14)
df = df.dropna(subset=['REVIEW']) # should consider only the 'REVIEW' column when removing rows with missing values
X= df['REVIEW']

predictions = svm_model.predict(X)

mapping = {1: "REAL", 0: "FAKE"}
start_row = 18

for i, prediction in enumerate(predictions):
    label = mapping[prediction]
    cell = ws.cell(row=start_row + i, column=17)  
    cell.value = label
    if label == "FAKE":
        cell.font = Font(bold=True)
        cell.fill =  PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

workbook.save("product_detected_reviews.xlsx")
workbook.close()


